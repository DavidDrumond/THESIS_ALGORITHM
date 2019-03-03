# -*- coding: utf-8 -*-
"""
Created on Thu Oct 04 10:25:10 2018

@author: David
"""

from __future__ import unicode_literals
import openpyxl as px 
import numpy as np 
import pandas as pd 
import sklearn.preprocessing as preprocessing 
import sklearn.decomposition as decomposition
from sklearn.model_selection import train_test_split 
import math
import matplotlib.pyplot as plt 
from sklearn import preprocessing
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.cluster import KMeans
from sklearn import svm
from sklearn.model_selection import cross_validate
from sklearn.preprocessing import QuantileTransformer
from sklearn.preprocessing import StandardScaler
import seaborn as sns
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.preprocessing import LabelEncoder 
from sklearn.linear_model import LogisticRegression
from pylab import *
import sklearn.discriminant_analysis as DA
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF
from sklearn.neural_network import MLPClassifier
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import chi2
from sklearn.feature_selection import RFE
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.model_selection import learning_curve
from sklearn.ensemble import AdaBoostClassifier
from sklearn.manifold import TSNE
from sklearn.metrics import classification_report
from pylab import pcolor, show, colorbar, xticks, yticks
from sklearn.metrics import confusion_matrix
import itertools
from sklearn import manifold
from sklearn.metrics import euclidean_distances
from scipy.cluster import hierarchy
from sklearn.ensemble import AdaBoostClassifier
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import accuracy_score
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import roc_curve, auc
from sklearn.preprocessing import label_binarize
from sklearn import tree
from sklearn.naive_bayes import BernoulliNB
import sklearn.metrics as m
import graphviz
from sklearn.tree import export_graphviz

def ReadData (endereco_dados, nome_sheet_dados, nome_sheet_variaveis):
    # READ DATA 
    print ("Iniciando carregamento dos dados")
    print(".............................................................................................")
    try:
        W = px.load_workbook(endereco_dados, data_only = True)
        p = W[nome_sheet_dados]
        p2 = W[nome_sheet_variaveis]    
        dados_planilha =[]
        variaveis =[]
        labels = []
        linha_inicial = 0
        skip_row = 0
        for row in p.iter_rows():
            b =[]
            linha_inicial += 1
            if linha_inicial ==  (skip_row +  1):
                for k in row:
                    labels.append(k.internal_value)
            if linha_inicial > (skip_row + 1): 
                for k in row:
                    b.append(k.internal_value)
                dados_planilha.append(b)
                
        dados_planilha = pd.DataFrame(dados_planilha, columns=labels)        
        for row in p2.iter_rows():
            for k in row:
                if k.value:
                    variaveis.append(k.internal_value)
    except:
        raise Exception("Endereco de dados ou configuração da planilha errada")
    print ("Carregamento dos dados concluido")
    print(".............................................................................................")
    return dados_planilha, variaveis


def Make_TSNE (Xnormalizer):
    tsne = TSNE(n_components = 2)
    X_PCA3 = tsne.fit_transform(Xnormalizer)
    df = pd.DataFrame(np.array([X_PCA3[:,0], X_PCA3[:,1], Ydata]).T, columns=["TSNE1","TSNE2", "Grupos"])
    sns.pairplot(x_vars=["TSNE1"], y_vars=["TSNE2"], data=df, hue='Grupos', size=5 )
    plt.title("TSNE Factors")
    plt.show()
    
def Make_UMAP (Xnormalizer):
    reducer = umap.UMAP().fit_transform(Xnormalizer)
    df = pd.DataFrame(np.array([reducer[:,0], reducer[:,1], Ydata]).T, columns=["UMAP1","UMAP2","Grupos"])
    um  = df["UMAP1"].astype(float)
    dois = df["UMAP2"].astype(float)
    print (um.corr(dois))
    sns.pairplot(x_vars=["UMAP1"], y_vars=["UMAP2"], data=df, hue="Grupos",size=5)
    plt.title("UMAP")
    plt.show()


def Prever_R_met(classificador, dados_planilha, X_PCA):
    # DETERMINAÇÃO DA RECUPERAÇÃO METALURGICA
    predicted = classificador.predict(X_PCA)
    ouro = dados_planilha['Au [ppm]']
    peso = dados_planilha['Weight']
    

    ouro_rejeito =[]
    ouro_alimentacao =[]
    ouro_concentrado = []
    massa_rejeito = 0
    massa_concentrado= 0
    massa_alimentacao = 0
    
   
    for j,i in enumerate(predicted):
        if i == 'MINERIO':
            if ouro[j] >= 0  and peso[j] > 0:
                ouro_concentrado.append(ouro[j]*peso[j])
                massa_concentrado += float(peso[j])
        if i == 'ESTERIL':
            if ouro[j] >= 0 and peso[j] > 0:
                ouro_rejeito.append(ouro[j]*float(peso[j]))
                massa_rejeito += peso[j]
        if ouro[j] >= 0 and peso[j]>0:
            ouro_alimentacao.append(ouro[j]*float(peso[j]))
            massa_alimentacao += peso[j]
                
   
    ouro_rejeito = np.array(ouro_rejeito)
    ouro_rejeito = ouro_rejeito[ouro_rejeito != None]

    ouro_concentrado = np.array(ouro_concentrado)
    ouro_concentrado= ouro_concentrado[ouro_concentrado!= None]

    ouro_alimentacao = np.array(ouro_alimentacao)
    ouro_alimentacao= ouro_alimentacao[ouro_alimentacao!= None]
    
    teor_concentrado = np.nansum(ouro_concentrado)/float(massa_concentrado)
    teor_alimentacao = np.nansum(ouro_alimentacao)/float(massa_alimentacao)
    print ("Recuperação metalúrgica do ouro no ore sorting")
    print (".............................................................................................")
    print (float(teor_concentrado*massa_concentrado)/float(teor_alimentacao*massa_alimentacao))

 

def Make_PCA_graph(nfactors, Xnormalizer, Ydata):
    #transform pca values
    pca = decomposition.PCA(n_components = nfactors)
    X_PCA2 = pca.fit_transform(Xnormalizer)
    X_PCA = pca.fit_transform(Xnormalizer)
    #X_PCA = Xnormalizer

    
    #make graph of pca values 
    df = pd.DataFrame(np.array([X_PCA2[:,0], X_PCA2[:,1], Ydata]).T, columns=["PCA1","PCA2", "Grupos"])
    sns.pairplot(x_vars=["PCA1"], y_vars=["PCA2"], data=df, hue='Grupos',markers ='+', size=7, plot_kws={'alpha': 0.8}  )
    plt.title("PCA Factors")
    plt.show()
    return X_PCA


def Remover_outliers (Ydata, Xdata, writer, ndesvpad):

    index_outliers =[]
    for i in range(len(Xdata)):
        for j in range (len(Xdata[0])):
            if Xdata[i][j] + np.mean(Xdata[i][j])> ndesvpad or Xdata[i][j] + np.mean(Xdata[i][j])  < -ndesvpad:
                index_outliers.append(i)
    
    index_outliers = list(set(index_outliers))
    
     # NORMALIZACAO DOS DADOS 
    #Xnormalizer = preprocessing.StandardScaler().fit(Xdata).transform(Xdata)
    Xnormalizer = np.array(Xdata)
    #Xnormalizer = QuantileTransformer(output_distribution='uniform').fit_transform(Xdata)
    indexes = range(0,len(Xnormalizer.T[0]))
    index_non_outliers = list(set(indexes) - set(index_outliers))
    
    # REMOVER VALORES OUTLIERS ACIMA OU ABAIXO DE 2 DESVPAD
    
    Xnormalizer = Xnormalizer[:][index_non_outliers]
    Ydata = Ydata[index_non_outliers]
    
    index_outliers = pd.DataFrame(index_outliers)
    index_outliers.to_excel(writer, 'Index outliers')
    
    return Xnormalizer, Ydata
    
def Importancia_variavies ( variaveis, Xnormalizer, Ydata, writer):
    
    # USE RANDOM FOREST CLASSIFIER TO DEFINE THE VARIABLE IMPORTANCE
    model = RandomForestClassifier(n_estimators=5000,n_jobs=4,max_features=3,max_depth=2)
    model.fit(Xnormalizer, Ydata)   
    
    variavel_importance = []
    for i, j in enumerate(model.feature_importances_):
        variavel_importance.append(j)   
    
    zipped = zip(variavel_importance, variaveis[1:])
    zipped = sorted(zipped, reverse =True)
    lista1 = [i for (i,s) in zipped]
    lista2 = [s for (i,s) in zipped]
    
    plt.figure(figsize=(19,8))
    X = np.arange(len(variavel_importance))
    plt.xticks(np.arange(len(variavel_importance)), lista2, rotation=20)
    plt.bar(X, lista1, color = 'b', width = 1, label="feature importance")
    plt.title("Importância das variáveis")
    plt.savefig(".\\IMAGENS\\importancia.png", dpi = 800)
    plt.show()   
    
def Create_ROC_curve(X_PCA, Ydata, model):

    lb = label_binarize(Ydata, classes=['MINERIO', 'ESTERIL'])
    fpr, tpr, _ = roc_curve(lb[:,0],model.fit(X_PCA, lb[:,0]).predict_proba(X_PCA)[:,1])
    

    plt.plot(fpr,tpr, label = "curva ROC")
    plt.plot([0, 1], [0, 1], color='navy',lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('TAXA DE FALSOS POSITIVOS')
    plt.ylabel('TAXA DE POSITIVOS VERDADEIROS')
    plt.title('ROC - Receiver operating characteristic ')
    plt.legend(loc="lower right")
    plt.show()
    
def Create_metrics_relatory(list_of_classificators,list_of_labels, X_PCA, Ydata):
    
        arquivo_saida =[]
        arquivo = open("relatorio_classificacao.txt","w")
        arquivo.write('modelo accuracy_mean accuracy_var f1-score_mean f1_score_var precision_mean precision_var recall_mean recall_var area_above_roc_mean area_above_roc_var \n' )
        kfold = StratifiedKFold(n_splits = 10, shuffle = True, random_state = 0)
        for modelo, nome in zip(list_of_classificators, list_of_labels):
            resultados_a = []
            resultados_f = []
            resultados_s = []
            resultados_r = []
            resultados_roc = []
            Ydata2 = np.array([1 if x == "MINERIO" else 0 for x in Ydata])
            for indice_treinamento, indice_teste in kfold.split(X_PCA,Ydata2):
                modelo.fit( X_PCA[indice_treinamento], Ydata2[indice_treinamento])
                previsoes = modelo.predict(X_PCA[indice_teste])
                resultados_a.append(m.accuracy_score(Ydata2[indice_teste], previsoes))
                resultados_f.append(m.f1_score(Ydata2[indice_teste], previsoes))
                resultados_s.append(m.precision_score(Ydata2[indice_teste], previsoes))
                resultados_r.append(m.recall_score(Ydata2[indice_teste], previsoes))
                resultados_roc.append(m.roc_auc_score(Ydata2[indice_teste], previsoes))
            resultados = np.asarray(resultados_a)
            resultados1 = np.asarray(resultados_f)
            resultados2 = np.asarray(resultados_s)
            resultados3 = np.asarray(resultados_r)
            resultados4 = np.asarray(resultados_roc)
            arquivo.write("{}  {} {} {} {} {} {} {} {} {} {} \n".format(nome,
                          str(resultados.mean()),str(resultados.var()),
                          str(resultados1.mean()),str(resultados1.var()),
                          str(resultados2.mean()),str(resultados2.var()),
                          str(resultados3.mean()),str(resultados3.var()),
                          str(resultados4.mean()),str(resultados4.var())))
        arquivo.close()



def PCA_importancia (labels_input, Xnormalizer, writer):
    pca = decomposition.PCA(n_components=len(labels_input))
    PCA = pca.fit_transform(Xnormalizer)
    factors_importance = np.array(pca.explained_variance_ratio_)
    factors_importance = sorted(factors_importance, reverse = True)
    accumulative_importance = np.cumsum(factors_importance)
    dados = np.stack((factors_importance,accumulative_importance), axis=-1)
    factors_importance = pd.DataFrame(dados, columns=['Factors Importance', 'Acumulative Importance'])
    factors_importance.to_excel(writer, 'Importancia_fatores') 
    print ("IMPORTANCIA PCA")
    print (".............................................................................................")
    print ([[round(n, 5) for n in accumulative_importance]])
    return accumulative_importance

def MDS (data, Ydata):
    

    data = np.array(data)
    similarities = euclidean_distances(data.T)
    mds = manifold.MDS(n_components=2,dissimilarity="precomputed")
    pos = mds.fit(similarities).embedding_   
    plt.figure(figsize=(10,10))
    plt.scatter(pos[:,0],pos[:,1])   
    for i, d in enumerate(Ydata):
        plt.annotate(d, (pos[:,0][i], pos[:,1][i]))
    plt.savefig(".\\IMAGENS\\MDS.png", dpi =500)
    plt.show()
    return 0

    


def Curva_Aprendizado(Xnormalizer, Ydata, model, titulo):
    # REALIZE A CURVA DE APRENDIZADO PARA OS MODELOS 
    #........................................................................
    
    train_sizes = [0.4,0.5,0.6,0.7,0.8,0.9,1]
    test =[]
    train =[]
    number_data =[]
    for i in train_sizes:
        X_train, X_test, Y_train, Y_test = train_test_split(Xnormalizer, Ydata, test_size=i, random_state=42)
        scores = cross_validate(model, X_train, Y_train, cv=5)
        test.append(np.mean(np.array(scores['test_score'])))
        train.append(np.mean(np.array(scores['train_score'])))
        number_data.append(i)
    
    fig, axes = plt.subplots()
    axes.plot(number_data, test, label="Média dos scores da cross validação - Teste")
    axes.plot(number_data, train, label= "Média dos scores da cross validação - Treino")
    axes.set_xlabel("Porcentagem dos dados de treino  "+titulo)
    axes.set_ylabel("Acurácia")
    plt.savefig(".\\IMAGENS\\curva_aprendizado{}.png".format(titulo), dpi =500)
    fig.legend()
    plt.show()
    
def classifaction_report_csv(report,writer,nome):
    # TRANSFORME O REPORT DE CLASSIFICAÇÃO EM ARQUIVO EXCEL 
    #..............................................................................
    report_data = []
    lines = report.split('\n')
    for line in lines[2:-3]:
        row = {}
        row_data = line.split()
        row['class'] = row_data[0]
        row['precision'] = float(row_data[1])
        row['recall'] = float(row_data[2])
        row['f1_score'] = float(row_data[3])
        row['support'] = float(row_data[4])
        report_data.append(row)
    dataframe = pd.DataFrame.from_dict(report_data)
    dataframe.to_excel(writer,nome)

def Classificators(X_PCA, Ydata, list_of_labels):
    
    list_of_classificators = {}
    for i in list_of_labels:
        if i not in ["LOGISTIC", "LDA", "SVM", "RF", "NEURAL", "KNC", "DT", "NB"]:
            raise Exception("There is no classification model with this label") 
    if "LOGISTIC" in list_of_labels:
        LOGISTIC = LogisticRegression ()
        LOGISTIC.fit(X_PCA, Ydata)
        list_of_classificators['LOGISTIC'] = LOGISTIC
    if "LDA" in list_of_labels:
        LDA = LinearDiscriminantAnalysis ()
        LDA.fit(X_PCA, Ydata)
        list_of_classificators['LDA'] = LDA
    if "SVM" in list_of_labels:
        SVM = svm.SVC(probability=True)
        SVM.fit(X_PCA, Ydata)
        list_of_classificators['SVM'] = SVM
    if "RF" in list_of_labels:
        RF = RandomForestClassifier(max_depth =4, criterion ='entropy', min_samples_split= 15, random_state =0)
        RF.fit(X_PCA, Ydata)
        list_of_classificators['RF'] = RF
    if "DT" in list_of_labels:
        DT = tree.DecisionTreeClassifier(max_depth =4, criterion ='entropy', min_samples_split= 15)
        DT.fit(X_PCA, Ydata)
        list_of_classificators['DT'] = DT
    if "NEURAL" in list_of_labels:
        NEURAL = MLPClassifier( solver='lbfgs', alpha=1e-5,hidden_layer_sizes=(5, 10), random_state=0)
        NEURAL.fit(X_PCA, Ydata)
        list_of_classificators['NEURAL'] = NEURAL
    if "KNC" in list_of_labels:
        KNC = KNeighborsClassifier(n_neighbors= 30)
        KNC.fit(X_PCA, Ydata)
        list_of_classificators['KNC'] = KNC
    if "NB" in list_of_labels:
        NB = BernoulliNB(alpha = 1.0)
        NB.fit(X_PCA, Ydata)
        list_of_classificators['NB'] = NB
    return list_of_classificators
    



def Confusion_Matrix(Xnormalizer, Ydata, modelo, nome, minerio_esteril ):
   
    kfold = StratifiedKFold(n_splits = 10, shuffle = True, random_state = 0)
    resultados = []
    
   

    # TRANSFORME VARIÁVEIS EM ESTÉRIL E  MINÉRIO OU VARIÁVEIS LITOTIPO
    #.........................................................................
    if minerio_esteril:
        Ydata2 = np.array([1 if x == "MINERIO" else 0 for x in Ydata])
        for indice_treinamento, indice_teste in kfold.split(Xnormalizer,Ydata2):
            modelo.fit( Xnormalizer[indice_treinamento], Ydata2[indice_treinamento])
            previsoes = modelo.predict(Xnormalizer[indice_teste])
            precisao = confusion_matrix(Ydata2[indice_teste], previsoes)
            resultados.append(precisao)
        resultados = np.asarray(resultados)
        classes = ["ESTÉRIL","MINÉRIO"]
    else:
        le = preprocessing.LabelEncoder()
        le.fit(Ydata)
        Ydata2 = le.transform(Ydata)
        classes = le.classes_
        for indice_treinamento, indice_teste in kfold.split(Xnormalizer,Ydata2):
            modelo.fit( Xnormalizer[indice_treinamento], Ydata2[indice_treinamento])
            previsoes = modelo.predict(Xnormalizer[indice_teste])
            precisao = confusion_matrix(Ydata2[indice_teste], previsoes)
            resultados.append(precisao)
        resultados = np.asarray(resultados)


    # FAÇA A MATRIZ DE CONFUSÃO
    #.....................................................................
    
    report = resultados.mean(axis = 0)
    report= report.astype('float') / report.sum(axis=1)[:, np.newaxis]
    np.set_printoptions(precision=1)
    df = pd.DataFrame(report, columns = classes)
    cmap = cm.get_cmap('Blues')
    plt.title(nome)
    sns.heatmap(df, 
        xticklabels=df.columns,
        yticklabels=df.columns, cmap=cmap,annot=True, fmt=".2f")
    plt.ylabel('Valor verdadeiro')
    plt.xlabel('Valor predito')
    plt.savefig(".\\IMAGENS\\confuse_matriz{}.png".format(nome), dpi = 500)
    plt.show()
    

    
   

def correlation_matrix(df, Variaveis):
    corr = df.corr()
    cmap = cm.get_cmap('Blues')
    plt.figure(figsize=(12,12))
    sns.heatmap(corr, 
        xticklabels=corr.columns,
        yticklabels=corr.columns, cmap=cmap,annot=True, fmt=".2f")
    plt.savefig(".\\IMAGENS\\correlation.png", dpi = 500)
    plt.show()
    
def histogram (df, output, size):
    
    colors =["skyblue","olive","gold","teal"]
    NROWS = int(size/2)
    NCOLS = int(size/NROWS)
    print (NROWS,NCOLS)
    o = 0
    plt.subplots(figsize=(10,10))
    for o in range(len(output)):
        plt.subplot(NROWS,NCOLS,(o+1))
        plt.hist(df[output[o]] , color=np.random.choice(colors))
        plt.xlabel(output[o])
        o +=1 
    plt.savefig(".\\IMAGENS\\histogram.png", dpi = 500)
    plt.show()


def Curva_concentracao (df, variaveis, minimo, maximo,delta):
    teor_concentrado = []
    concentracao = []
    alimentacao = df[df[variaveis[0]]== 'MINERIO'].shape[0]/float(df[df[variaveis[0]] == 'ESTERIL'].shape[0]+df[df[variaveis[0]]== 'MINERIO'].shape[0])
    for i in np.linspace(minimo, maximo,delta):
        acima = df[(df[variaveis[1]]) > i]
        abaixo = df[(df[variaveis[1]]) < i]
        minerio_conc = acima[(acima[variaveis[0]] == 'MINERIO')].shape[0]
        esteril_conc = acima[(acima[variaveis[0]] == 'ESTERIL')].shape[0]
        minerio_rej = abaixo[(abaixo[variaveis[0]] == 'MINERIO')].shape[0]
        esteril_rej = abaixo[(abaixo[variaveis[0]] == 'ESTERIL')].shape[0]
        if (minerio_conc+esteril_conc) > 0:
            concentrado = minerio_conc/ float(minerio_conc +esteril_conc)
        else:
            concentrado = 0
        if (minerio_rej +esteril_rej) > 0:
            rejeito = minerio_rej/ float(minerio_rej +esteril_rej)
        else:
            rejeito = 0
        teor_concentrado.append(concentrado*100)

        concentracao.append(concentrado/alimentacao*(alimentacao-rejeito)/(concentrado-rejeito)*100)
    
    plt.title("Curva teor concentrado e recuperação por variável")
    plt.xlabel(variaveis[1])
    plt.ylabel("%")
    plt.plot(np.linspace(minimo,maximo,delta), teor_concentrado, label = "Concentração")
    plt.plot(np.linspace(minimo,maximo,delta), concentracao, label="Recuperação")
    plt.legend()
    plt.savefig(".\\IMAGENS\\recuperacao.png", dpi = 500)
    plt.show()
    
def meshgrid2(*arrs):
    arrs = tuple(reversed(arrs))  #edit
    lens = map(len, arrs)
    dim = len(arrs)

    sz = 1
    for s in lens:
        sz*=s

    ans = []    
    for i, arr in enumerate(arrs):
        slc = [1]*dim
        slc[i] = lens[i]
        arr2 = asarray(arr).reshape(slc)
        for j, sz in enumerate(lens):
            if j!=i:
                arr2 = arr2.repeat(sz, axis=j) 
        ans.append(arr2)

    return tuple(ans)

def Create_classification_charts(X_PCA, Ydata, classificators, list_of_labels, minerio_esteril, caprend = True, cconf = True):
    for modelo, nome in zip(classificators, list_of_labels):
        if caprend == True:
            Curva_Aprendizado(X_PCA, Ydata, modelo, nome)
        if cconf == True:
            Confusion_Matrix(X_PCA, Ydata, modelo, nome,minerio_esteril )
        #Make_image(X_PCA,Ydata,nome,modelo, pca, Xnormalizer, variaveis,minerio_esteril)
    
    
def Make_image(X_PCA,Ydata,nome,modelo, pca, Xnormalizer, variaveis,minerio_esteril):
        h = 1
        plt.figure(figsize=(8, 6))
        x_min, x_max = X_PCA[:,0].min() - h, X_PCA[:,0].max() + h
        y_min, y_max = X_PCA[:,1].min() - h, X_PCA[:,1].max() + h
        data = np.mgrid[x_min:x_max:200j,y_min:y_max:200j]
        x_p = np.linspace(x_min,x_max,200)
        y_p = np.linspace(y_min,y_max,200)

        data = np.vstack(np.meshgrid(x_p,y_p)).reshape(2,-1).T
        data_inv = pca.inverse_transform(data)

        
        # ORIGINAL DATA IMAGEM
        Z = modelo.predict(data)
        
        if minerio_esteril == True:
            Z = np.array([1 if x=='MINERIO' else 0 for x in Z])
            Y = np.array([1 if x=='MINERIO' else 0 for x in Ydata])
        else:
            le = preprocessing.LabelEncoder()
            le.fit(Ydata)
            Y = le.transform(Ydata)
            Z = le.transform(Z)
        

        if minerio_esteril == True:
            plt.scatter(data_inv[:,0],data_inv[:,1],c=Z, cmap =plt.cm.coolwarm, alpha = 0.3)
            plt.scatter(Xnormalizer[:,0],Xnormalizer[:,1],c =Y, cmap= plt.cm.coolwarm,edgecolors='k')
            plt.xlim(min(Xnormalizer[:,0]),max(Xnormalizer[:,0]))
            plt.ylim(min(Xnormalizer[:,1]),max(Xnormalizer[:,1]))
            cbar = plt.colorbar(ticks=[0,1])
            cbar.ax.set_yticklabels(['ESTERIL','MINERIO'])
        else: 
            plt.scatter(data_inv[:,0],data_inv[:,1],c=Z, cmap =plt.cm.tab20c, alpha = 0.3)
            plt.scatter(Xnormalizer[:,0],Xnormalizer[:,1],c =Y, cmap= plt.cm.coolwarm,edgecolors='k')
            plt.xlim(min(Xnormalizer[:,0]),max(Xnormalizer[:,0]))
            plt.ylim(min(Xnormalizer[:,1]),max(Xnormalizer[:,1]))
            x = range(0,len(le.classes_))
            cbar = plt.colorbar(ticks= x)
            cbar.ax.set_yticklabels(le.classes_)
        plt.title(("MODELO {} variavel {} x variavel {}").format(nome, variaveis[1], variaveis[2]))
        plt.xlabel("{}".format(variaveis[1]))
        plt.ylabel("{}".format(variaveis[2]))
        plt.show()
        
        # PCA IMAGEM
        xx, yy = np.meshgrid(x_p,y_p)
        
        
        if minerio_esteril == True:
            Z = Z.reshape(xx.shape)
            Y = np.array([1 if x=='MINERIO' else 0 for x in Ydata])
        else:
            le = preprocessing.LabelEncoder()
            le.fit(Ydata)
            Y = le.transform(Ydata)
            Z = Z.reshape(xx.shape)
        
        plt.pcolormesh(xx, yy, Z, cmap=plt.cm.coolwarm,alpha=0.3)
        plt.scatter(X_PCA[:,0], X_PCA[:,1], c=Y, cmap=plt.cm.coolwarm, s=20, edgecolors='k')
        #plt.yscale('log',basey=2)
        #plt.xscale('log',basex =2)
        plt.xlim(x_min+h,x_max-h)
        plt.ylim(y_min+h,y_max-h)
        
        if minerio_esteril == True:
            cbar = plt.colorbar(ticks=[0,1])
            cbar.ax.set_yticklabels(['ESTERIL','MINERIO'])
        else: 
            x = range(0,len(le.classes_))
            cbar = plt.colorbar(ticks= x)
            cbar.ax.set_yticklabels(le.classes_)
        
        plt.title(("MODELO {} PCA_1 X PCA_2").format(nome))
        plt.xlabel("1st eigenvector")
        plt.ylabel("2nd eigenvector")
        plt.savefig(".\IMAGENS\imagen_modelo{}.png".format(nome), dpi = 500)
        plt.show()
       
def print_contagem(Ydata):
    df = pd.DataFrame((Ydata).T,columns=['Grupos'])
    sns.set()
    fig = plt.figure(figsize=(8,4))
    sns.countplot(x='Grupos', data=df, palette="Greens_d") 
    plt.xlabel("Grupos")
    plt.ylabel("Contagem")
    plt.savefig(".\IMAGENS\count.png", dpi = 500)
    plt.show(fig)
    
def print_dados(Xnormalizer, Ydata,labels_input,labels):
    # REALIZAR GRÁFICO DOS DADOS
    x_label = labels_input[1]
    y_label = labels_input[0]

    df = pd.DataFrame(np.array([Xnormalizer[:,1], Xnormalizer[:,0], Ydata]).T,columns=[x_label,y_label, 'Grupos'])
    plt.figure(figsize=(20,15))
    sns.pairplot(x_vars=[x_label], y_vars=[y_label], data=df, hue='Grupos',markers ='+', size=7, plot_kws={'alpha': 0.8} )
    #plt.xscale('log')
    #plt.yscale('log')
    plt.show()
    sns.set()
    plt.figure(figsize=(8,4))
    sns.countplot(x='Grupos', data=df, palette="Greens_d") 
    plt.savefig(".\IMAGENS\dados.png", dpi = 500)
    plt.show()
    

    
    
    
    
    
    
    
    