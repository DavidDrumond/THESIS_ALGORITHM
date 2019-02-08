# -*- coding: utf-8 -*-
"""
Created on Thu Oct 04 10:25:10 2018

@author: David
"""

from __future__ import unicode_literals
import openpyxl as px 
import numpy as np 
import pandas as pd 
import sklearn.preprocessing as preprocessing 
import sklearn.decomposition as decomposition
from sklearn.model_selection import train_test_split 
import math
import matplotlib.pyplot as plt 
from sklearn import preprocessing
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.cluster import KMeans
from sklearn import svm
from sklearn.model_selection import cross_validate
from sklearn.preprocessing import QuantileTransformer
from sklearn.preprocessing import StandardScaler
import seaborn as sns
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.preprocessing import LabelEncoder 
from sklearn.linear_model import LogisticRegression
from pylab import *
import  sklearn.discriminant_analysis as DA
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF
from sklearn.neural_network import MLPClassifier
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import chi2
from sklearn.feature_selection import RFE
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.model_selection import learning_curve
from sklearn.ensemble import AdaBoostClassifier
from sklearn.manifold import TSNE
from sklearn.metrics import classification_report
from pylab import pcolor, show, colorbar, xticks, yticks
from sklearn.metrics import confusion_matrix
import itertools
from sklearn import manifold
from sklearn.metrics import euclidean_distances
from scipy.cluster import hierarchy
from sklearn.ensemble import AdaBoostClassifier
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import accuracy_score

def ReadData (endereco_dados, nome_sheet_dados, nome_sheet_variaveis):
    # READ DATA 
    print "Iniciando load dos dados"
    W = px.load_workbook(endereco_dados, data_only = True)
    print "step 1 ok"
    p = W[nome_sheet_dados]
    p2 = W[nome_sheet_variaveis]    
    dados_planilha =[]
    variaveis =[]
    labels = []
    linha_inicial = 0
    skip_row = 0
    for row in p.iter_rows():
        b =[]
        linha_inicial += 1
        if linha_inicial ==  (skip_row +  1):
            for k in row:
                labels.append(k.internal_value)
        if linha_inicial > (skip_row + 1): 
            for k in row:
                b.append(k.internal_value)
            dados_planilha.append(b)
            
    dados_planilha = pd.DataFrame(dados_planilha, columns=labels)        
    for row in p2.iter_rows():
        for k in row:
            if k.value:
                variaveis.append(k.internal_value)
    return dados_planilha, variaveis

def Remover_outliers (Ydata, Xdata, writer, ndesvpad):

    index_outliers =[]
    for i in xrange(len(Xdata)):
        for j in xrange (len(Xdata[0])):
            if Xdata[i][j] + np.mean(Xdata[i][j])> ndesvpad or Xdata[i][j] + np.mean(Xdata[i][j])  < -ndesvpad:
                index_outliers.append(i)
    
    index_outliers = list(set(index_outliers))
    
     # NORMALIZACAO DOS DADOS 
    #Xnormalizer = preprocessing.StandardScaler().fit(Xdata).transform(Xdata)
    Xnormalizer = np.array(Xdata)
    #Xnormalizer = QuantileTransformer(output_distribution='uniform').fit_transform(Xdata)
    indexes = range(0,len(Xnormalizer.T[0]))
    index_non_outliers = list(set(indexes) - set(index_outliers))
    
    # REMOVER VALORES OUTLIERS ACIMA OU ABAIXO DE 2 DESVPAD
    
    Xnormalizer = Xnormalizer[:][index_non_outliers]
    Ydata = Ydata[index_non_outliers]
    
    index_outliers = pd.DataFrame(index_outliers)
    index_outliers.to_excel(writer, 'Index outliers')
    
    return Xnormalizer, Ydata
    
def Importancia_variavies ( variaveis, Xnormalizer, Ydata, writer):
    
    # USE RANDOM FOREST CLASSIFIER TO DEFINE THE VARIABLE IMPORTANCE
    model = RandomForestClassifier(n_estimators=5000,n_jobs=4,max_features=3,max_depth=2)
    model.fit(Xnormalizer, Ydata)   
    
    variavel_importance = []
    for i, j in enumerate(model.feature_importances_):
        variavel_importance.append(j)   
    
    zipped = zip(variavel_importance, variaveis[1:])
    zipped = sorted(zipped, reverse =True)
    lista1 = [i for (i,s) in zipped]
    lista2 = [s for (i,s) in zipped]
    
    plt.figure(figsize=(19,8))
    X = np.arange(len(variavel_importance))
    plt.xticks(np.arange(len(variavel_importance)), lista2, rotation=20)
    plt.bar(X, lista1, color = 'b', width = 1, label="feature importance")
    plt.title("Importância das variáveis")
    plt.savefig(".\\IMAGENS\\importancia.png", dpi = 800)
    plt.show()   
    
    
    
    
def PCA_importancia (labels_input, Xnormalizer, writer):
    pca = decomposition.PCA(n_components=len(labels_input))
    PCA = pca.fit_transform(Xnormalizer)
    factors_importance = np.array(pca.explained_variance_ratio_)
    factors_importance = sorted(factors_importance, reverse = True)
    accumulative_importance = np.cumsum(factors_importance)
    dados = np.stack((factors_importance,accumulative_importance), axis=-1)
    factors_importance = pd.DataFrame(dados, columns=['Factors Importance', 'Acumulative Importance'])
    factors_importance.to_excel(writer, 'Importancia_fatores')   
    return accumulative_importance

def HCA (data, Ydata):
    
    data = np.array(data)
    similarities = euclidean_distances(data.T)
    mds = manifold.MDS(n_components=2,dissimilarity="precomputed")
    pos = mds.fit(similarities).embedding_   
    plt.figure(figsize=(30,30))
    plt.scatter(pos[:,0],pos[:,1])   
    for i, d in enumerate(Ydata):
        plt.annotate(d, (pos[:,0][i], pos[:,1][i]))
    plt.show()
    return 0

    


def Curva_Aprendizado(Xnormalizer, Ydata, model, titulo):
    # REALIZE A CURVA DE APRENDIZADO PARA OS MODELOS 
    #........................................................................
    
    train_sizes = [0.4,0.5,0.6,0.7,0.8,0.9,1]
    test =[]
    train =[]
    number_data =[]
    for i in train_sizes:
        X_train, X_test, Y_train, Y_test = train_test_split(Xnormalizer, Ydata, test_size=i, random_state=42)
        scores = cross_validate(model, X_train, Y_train, cv=5)
        test.append(np.mean(np.array(scores['test_score'])))
        train.append(np.mean(np.array(scores['train_score'])))
        number_data.append(i)
    
    fig, axes = plt.subplots()
    axes.plot(number_data, test, label="Média dos scores da cross validação - Teste")
    axes.plot(number_data, train, label= "Média dos scores da cross validação - Treino")
    axes.set_xlabel("Porcentagem dos dados de treino"+titulo)
    axes.set_ylabel("Score")
    plt.savefig(".\\IMAGENS\\curva_aprendizado{}.png".format(titulo), dpi =800)
    fig.legend()
    plt.show()
    
def classifaction_report_csv(report,writer,nome):
    # TRANSFORME O REPORT DE CLASSIFICAÇÃO EM ARQUIVO EXCEL 
    #..............................................................................
    report_data = []
    lines = report.split('\n')
    for line in lines[2:-3]:
        row = {}
        row_data = line.split()
        row['class'] = row_data[0]
        row['precision'] = float(row_data[1])
        row['recall'] = float(row_data[2])
        row['f1_score'] = float(row_data[3])
        row['support'] = float(row_data[4])
        report_data.append(row)
    dataframe = pd.DataFrame.from_dict(report_data)
    dataframe.to_excel(writer,nome)

def Report_Classificacao(Xnormalizer, Ydata, modelo, nome,writer, minerio_esteril ):
   
    kfold = StratifiedKFold(n_splits = 10, shuffle = True, random_state = 0)
    resultados = []
    
   

    # TRANSFORME VARIÁVEIS EM ESTÉRIL E  MINÉRIO OU VARIÁVEIS LITOTIPO
    #.........................................................................
    if minerio_esteril:
        Ydata2 = np.array([1 if x == "MINERIO" else 0 for x in Ydata])
        for indice_treinamento, indice_teste in kfold.split(Xnormalizer,Ydata2):
            modelo.fit( Xnormalizer[indice_treinamento], Ydata2[indice_treinamento])
            previsoes = modelo.predict(Xnormalizer[indice_teste])
            precisao = confusion_matrix(Ydata2[indice_teste], previsoes)
            resultados.append(precisao)
        resultados = np.asarray(resultados)
        classes = ["ESTERIL","MINERIO"]
    else:
        le = preprocessing.LabelEncoder()
        le.fit(Ydata)
        Ydata2 = le.transform(Ydata)
        classes = le.classes_
        for indice_treinamento, indice_teste in kfold.split(Xnormalizer,Ydata2):
            modelo.fit( Xnormalizer[indice_treinamento], Ydata2[indice_treinamento])
            previsoes = modelo.predict(Xnormalizer[indice_teste])
            precisao = confusion_matrix(Ydata2[indice_teste], previsoes)
            resultados.append(precisao)
        resultados = np.asarray(resultados)


    # FAÇA A MATRIZ DE CONFUSÃO
    #.....................................................................
    
    report = resultados.mean(axis = 0)
    report= report.astype('float') / report.sum(axis=1)[:, np.newaxis]
    np.set_printoptions(precision=1)
    df = pd.DataFrame(report, columns = classes)
    cmap = cm.get_cmap('Blues')
    plt.title(nome)
    sns.heatmap(df, 
        xticklabels=df.columns,
        yticklabels=df.columns, cmap=cmap,annot=True, fmt=".2f")
    plt.ylabel('Valor verdadeiro')
    plt.xlabel('Valor predito')
    plt.savefig(".\\IMAGENS\\confuse_matriz{}.png".format(nome), dpi = 800)
    plt.show()
    df.to_excel(writer,nome)
    
    
    # PRINTAR O REPORTE DE CLASSIFICAÇÃO
    #........................................................................
    
    #classification_report_dados = classification_report(Y_test,predicted)
    #classifaction_report_csv(classification_report_dados,writer,nome)
    
   

def correlation_matrix(df, Variaveis):
    corr = df.corr()
    cmap = cm.get_cmap('Blues')
    plt.figure(figsize=(12,12))
    sns.heatmap(corr, 
        xticklabels=corr.columns,
        yticklabels=corr.columns, cmap=cmap,annot=True, fmt=".2f")
    plt.savefig(".\\IMAGENS\\correlation.png", dpi = 2000)
    plt.show()
    
def histogram (df, output):
    
    tamanho = df.shape[1]/2 + df.shape[1]%2

    if df.shape[1]%2 == 0 and df.shape[1] > 2:
        f, axes = plt.subplots(tamanho, 2, figsize=(12, 14), sharex=True)
        p = 0
        for i in xrange(tamanho):
            for j in xrange(2):
                sns.distplot( df[output[p]] , ax=axes[i,j])
                p += 1
    else: 
        f, axes = plt.subplots(df.shape[1], 1, figsize=(12, 14), sharex=True)
        for i in xrange(df.shape[1]):
            sns.distplot(df[output[i]], ax= axes[i])
    plt.savefig(".\\IMAGENS\\histogram.png", dpi = 2000)
    plt.show()

def Curva_concentracao (df, variaveis, minimo, maximo,delta):
    teor_concentrado = []
    concentracao = []
    alimentacao = df[df[variaveis[0]]== 'MINERIO'].shape[0]/float(df[df[variaveis[0]] == 'ESTERIL'].shape[0]+df[df[variaveis[0]]== 'MINERIO'].shape[0])
    for i in np.linspace(minimo, maximo,delta):
        acima = df[(df[variaveis[1]]) > i]
        abaixo = df[(df[variaveis[1]]) < i]
        minerio_conc = acima[(acima[variaveis[0]] == 'MINERIO')].shape[0]
        esteril_conc = acima[(acima[variaveis[0]] == 'ESTERIL')].shape[0]
        minerio_rej = abaixo[(abaixo[variaveis[0]] == 'MINERIO')].shape[0]
        esteril_rej = abaixo[(abaixo[variaveis[0]] == 'ESTERIL')].shape[0]
        if (minerio_conc+esteril_conc) > 0:
            concentrado = minerio_conc/ float(minerio_conc +esteril_conc)
        else:
            concentrado = 0
        if (minerio_rej +esteril_rej) > 0:
            rejeito = minerio_rej/ float(minerio_rej +esteril_rej)
        else:
            rejeito = 0
        teor_concentrado.append(concentrado*100)

        concentracao.append(concentrado/alimentacao*(alimentacao-rejeito)/(concentrado-rejeito)*100)
    
    plt.title("Curva teor concentrado e recuperação por variável")
    plt.xlabel(variaveis[1])
    plt.ylabel("%")
    plt.plot(np.linspace(minimo,maximo,delta), teor_concentrado, label = "Concentração")
    plt.plot(np.linspace(minimo,maximo,delta), concentracao, label="Recuperação")
    plt.legend()
    plt.savefig(".\\IMAGENS\\recuperacao.png", dpi = 800)
    plt.show()
    
def meshgrid2(*arrs):
    arrs = tuple(reversed(arrs))  #edit
    lens = map(len, arrs)
    dim = len(arrs)

    sz = 1
    for s in lens:
        sz*=s

    ans = []    
    for i, arr in enumerate(arrs):
        slc = [1]*dim
        slc[i] = lens[i]
        arr2 = asarray(arr).reshape(slc)
        for j, sz in enumerate(lens):
            if j!=i:
                arr2 = arr2.repeat(sz, axis=j) 
        ans.append(arr2)

    return tuple(ans)


    
    
def Make_image(X_PCA,Ydata,nome,modelo, pca, Xnormalizer, variaveis,minerio_esteril):
        h = 1
        plt.figure(figsize=(8, 6))
        x_min, x_max = X_PCA[:,0].min() - h, X_PCA[:,0].max() + h
        y_min, y_max = X_PCA[:,1].min() - h, X_PCA[:,1].max() + h
        data = np.mgrid[x_min:x_max:200j,y_min:y_max:200j]
        x_p = np.linspace(x_min,x_max,200)
        y_p = np.linspace(y_min,y_max,200)

        data = np.vstack(np.meshgrid(x_p,y_p)).reshape(2,-1).T
        data_inv = pca.inverse_transform(data)

        
        # ORIGINAL DATA IMAGEM
        Z = modelo.predict(data)
        
        if minerio_esteril == True:
            Z = np.array([1 if x=='MINERIO' else 0 for x in Z])
            Y = np.array([1 if x=='MINERIO' else 0 for x in Ydata])
        else:
            le = preprocessing.LabelEncoder()
            le.fit(Ydata)
            Y = le.transform(Ydata)
            Z = le.transform(Z)
        

        if minerio_esteril == True:
            plt.scatter(data_inv[:,0],data_inv[:,1],c=Z, cmap =plt.cm.coolwarm, alpha = 0.3)
            plt.scatter(Xnormalizer[:,0],Xnormalizer[:,1],c =Y, cmap= plt.cm.coolwarm,edgecolors='k')
            plt.xlim(min(Xnormalizer[:,0]),max(Xnormalizer[:,0]))
            plt.ylim(min(Xnormalizer[:,1]),max(Xnormalizer[:,1]))
            cbar = plt.colorbar(ticks=[0,1])
            cbar.ax.set_yticklabels(['ESTERIL','MINERIO'])
        else: 
            plt.scatter(data_inv[:,0],data_inv[:,1],c=Z, cmap =plt.cm.tab20c, alpha = 0.3)
            plt.scatter(Xnormalizer[:,0],Xnormalizer[:,1],c =Y, cmap= plt.cm.coolwarm,edgecolors='k')
            plt.xlim(min(Xnormalizer[:,0]),max(Xnormalizer[:,0]))
            plt.ylim(min(Xnormalizer[:,1]),max(Xnormalizer[:,1]))
            x = range(0,len(le.classes_))
            cbar = plt.colorbar(ticks= x)
            cbar.ax.set_yticklabels(le.classes_)
        plt.title(("MODELO {} variavel {} x variavel {}").format(nome, variaveis[1], variaveis[2]))
        plt.xlabel("{}".format(variaveis[1]))
        plt.ylabel("{}".format(variaveis[2]))
        plt.show()
        
        # PCA IMAGEM
        xx, yy = np.meshgrid(x_p,y_p)
        
        
        if minerio_esteril == True:
            Z = Z.reshape(xx.shape)
            Y = np.array([1 if x=='MINERIO' else 0 for x in Ydata])
        else:
            le = preprocessing.LabelEncoder()
            le.fit(Ydata)
            Y = le.transform(Ydata)
            Z = Z.reshape(xx.shape)
        
        plt.pcolormesh(xx, yy, Z, cmap=plt.cm.coolwarm,alpha=0.3)
        plt.scatter(X_PCA[:,0], X_PCA[:,1], c=Y, cmap=plt.cm.coolwarm, s=20, edgecolors='k')
        #plt.yscale('log',basey=2)
        #plt.xscale('log',basex =2)
        plt.xlim(x_min+h,x_max-h)
        plt.ylim(y_min+h,y_max-h)
        
        if minerio_esteril == True:
            cbar = plt.colorbar(ticks=[0,1])
            cbar.ax.set_yticklabels(['ESTERIL','MINERIO'])
        else: 
            x = range(0,len(le.classes_))
            cbar = plt.colorbar(ticks= x)
            cbar.ax.set_yticklabels(le.classes_)
        
        plt.title(("MODELO {} PCA_1 X PCA_2").format(nome))
        plt.xlabel("1st eigenvector")
        plt.ylabel("2nd eigenvector")
        plt.savefig(".\IMAGENS\imagen_modelo{}.png".format(nome), dpi = 800)
        plt.show()
       
def print_contagem(Ydata):
    df = pd.DataFrame((Ydata).T,columns=['Grupos'])
    sns.set()
    fig = plt.figure(figsize=(8,4))
    sns.countplot(x='Grupos', data=df, palette="Greens_d") 
    plt.xlabel("Grupos")
    plt.ylabel("Contagem")
    plt.savefig(".\IMAGENS\count.png", dpi = 800)
    plt.show(fig)
    
def print_dados(Xnormalizer, Ydata,labels_input,labels):
    # REALIZAR GRÁFICO DOS DADOS

    x_label = labels_input[1]
    y_label = labels_input[0]

    df = pd.DataFrame(np.array([Xnormalizer[:,1], Xnormalizer[:,0], Ydata]).T,columns=[x_label,y_label, 'Grupos'])
    plt.figure(figsize=(20,15))
    sns.pairplot(x_vars=[x_label], y_vars=[y_label], data=df, hue='Grupos',markers ='+', size=7, plot_kws={'alpha': 0.8} )
    #plt.xscale('log')
    #plt.yscale('log')
    plt.show()
    
    sns.set()
    plt.figure(figsize=(8,4))
    sns.countplot(x='Grupos', data=df, palette="Greens_d") 
    plt.savefig(".\IMAGENS\dados.png", dpi = 800)
    plt.show()
    
def Create_supersec(Xvalores, Ylabel):
    np.cov(Xvalores)
    
    
    
    
    
    
    
    